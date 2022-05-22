# built-in's
import os 
from datetime import datetime as dt

from pyparsing import col
# mi código
from stats.models import Activity
# externo
import openpyxl
from openpyxl.styles import Font

#Inicializa
dir_path = os.path.dirname(os.path.realpath(__file__))
PATH = fr'{dir_path}\sheets\sample.xlsx'
wb = openpyxl.Workbook()
sheet = wb['Sheet']
colNames = ['Registro','Fecha','Monto']

def create():
    data = getData()

    for (i,k) in enumerate(data):
        for j in range(3):
            sheet.cell(row=i+1,column=j+1).value = k[j]
        
    long = len(data)
    
    sheet.cell(row=long+1,column=3).value = f"= SUM(C1:C{long})"
    sheet.cell(row=long+1,column=3).font = Font(color='bd1616')
    wb.save(PATH)
    sheetName = dt.now().strftime('%Y-%m-%d')
    wb.save(f"{os.path.expanduser('~')}\\Downloads\\Sheet-{sheetName}.xlsx")
    wb.close()

def getData():

    month = dt.now().month
    data = Activity.objects.filter(date__month=month)
    data_list = list(data.values())

    table = [colNames]
    for k in data_list:
        k['date'] = str(k['date'])
        table.append(list(k.values())[1:4])
        
    return table




